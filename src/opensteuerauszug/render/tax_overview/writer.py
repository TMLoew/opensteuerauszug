"""Tax-overview writer: orchestrates xlsx / html / pdf output.

Phase 1 produces a minimal but valid workbook skeleton (Übersicht sheet only,
with placeholder content). Later phases add sheets, HTML, and PDF.
"""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Literal, Optional

from openpyxl import Workbook

from .design import StyleName, register_named_styles


OutputFormat = Literal["xlsx", "html", "pdf"]
ALL_FORMATS: tuple[OutputFormat, ...] = ("xlsx", "html", "pdf")


@dataclass(frozen=True)
class TaxOverviewRequest:
    """Inputs for one tax-overview invocation.

    Kept intentionally small: the CLI layer resolves filesystem paths and
    importer types; this request carries only what the writer needs.
    """

    input_path: Path
    broker: str
    tax_year: int
    output_dir: Path
    formats: tuple[OutputFormat, ...]
    preparer_mode: bool


def compute_report_hash(request: TaxOverviewRequest) -> str:
    """Deterministic hash derived from input data only.

    Stable across preparer and third-party exports: the same input file + year
    produces the same hash regardless of which output variants are requested.
    """
    h = hashlib.sha256()
    h.update(str(request.tax_year).encode("utf-8"))
    h.update(b"\x00")
    h.update(request.broker.encode("utf-8"))
    h.update(b"\x00")
    with request.input_path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def write_tax_overview(request: TaxOverviewRequest) -> List[Path]:
    """Entry point. Produce the requested formats, return the generated paths."""
    request.output_dir.mkdir(parents=True, exist_ok=True)
    report_hash = compute_report_hash(request)

    produced: List[Path] = []
    for fmt in request.formats:
        if fmt == "xlsx":
            produced.append(_write_xlsx(request, report_hash))
        elif fmt == "html":
            produced.append(_write_html_stub(request, report_hash))
        elif fmt == "pdf":
            produced.append(_write_pdf_stub(request, report_hash))
        else:  # pragma: no cover - typer validates enum
            raise ValueError(f"unknown output format: {fmt}")
    return produced


def _output_path(request: TaxOverviewRequest, suffix: str) -> Path:
    return request.output_dir / f"tax_overview_{request.tax_year}{suffix}"


def _write_xlsx(request: TaxOverviewRequest, report_hash: str) -> Path:
    wb = Workbook()
    register_named_styles(wb)

    ws = wb.active
    ws.title = "Übersicht"
    ws["A1"] = f"Steuer-Übersicht {request.tax_year}"
    ws["A1"].style = StyleName.KPI_VALUE
    ws["A3"] = "Platzhalter — Inhalt folgt in nachfolgenden Phasen."
    ws["A3"].style = StyleName.BODY_TEXT
    ws["A5"] = "Belegnummer"
    ws["A5"].style = StyleName.KPI_LABEL
    ws["B5"] = report_hash[:8]
    ws["B5"].style = StyleName.BODY_TEXT
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40

    path = _output_path(request, ".xlsx")
    wb.save(path)
    return path


def _write_html_stub(request: TaxOverviewRequest, report_hash: str) -> Path:
    path = _output_path(request, ".html")
    path.write_text(
        f"<!doctype html>\n"
        f"<html lang=\"de-CH\"><head><meta charset=\"utf-8\">"
        f"<title>Steuer-Übersicht {request.tax_year}</title></head>"
        f"<body><h1>Steuer-Übersicht {request.tax_year}</h1>"
        f"<p>Platzhalter — Inhalt folgt.</p>"
        f"<p>Belegnummer: {report_hash[:8]}</p></body></html>\n",
        encoding="utf-8",
    )
    return path


def _write_pdf_stub(request: TaxOverviewRequest, report_hash: str) -> Path:
    # Produce a tiny valid PDF without pulling in reportlab yet — phase 9
    # replaces this with a proper cover page.
    path = _output_path(request, ".pdf")
    pdf_bytes = _minimal_pdf_bytes(
        f"Steuer-Übersicht {request.tax_year} — Beleg {report_hash[:8]}"
    )
    path.write_bytes(pdf_bytes)
    return path


def _minimal_pdf_bytes(text: str) -> bytes:
    """Smallest valid single-page PDF containing the given text.

    Hand-rolled so the phase-1 stub has no dependency on reportlab's formatting
    machinery. Phase 9 replaces this with the real cover writer.
    """
    safe = text.replace("(", r"\(").replace(")", r"\)")
    stream = f"BT /F1 18 Tf 72 720 Td ({safe}) Tj ET".encode("latin-1", errors="replace")
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        (b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
         b"/Resources << /Font << /F1 5 0 R >> >> /Contents 4 0 R >>"),
        b"<< /Length " + str(len(stream)).encode() + b" >>\nstream\n" + stream + b"\nendstream",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]
    out = bytearray(b"%PDF-1.4\n")
    offsets: list[int] = []
    for idx, obj in enumerate(objects, start=1):
        offsets.append(len(out))
        out += f"{idx} 0 obj\n".encode() + obj + b"\nendobj\n"
    xref_pos = len(out)
    out += f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n".encode()
    for off in offsets:
        out += f"{off:010d} 00000 n \n".encode()
    out += (
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
        f"startxref\n{xref_pos}\n%%EOF\n"
    ).encode()
    return bytes(out)


def parse_formats(value: Optional[str]) -> tuple[OutputFormat, ...]:
    """Parse the --formats CLI value. None or empty means all formats."""
    if value is None or value.strip() == "":
        return ALL_FORMATS
    requested = [p.strip().lower() for p in value.split(",") if p.strip()]
    unknown = [f for f in requested if f not in ALL_FORMATS]
    if unknown:
        raise ValueError(
            f"unknown output format(s): {', '.join(unknown)}. "
            f"Known: {', '.join(ALL_FORMATS)}"
        )
    # Preserve canonical order.
    return tuple(f for f in ALL_FORMATS if f in requested)  # type: ignore[misc]
