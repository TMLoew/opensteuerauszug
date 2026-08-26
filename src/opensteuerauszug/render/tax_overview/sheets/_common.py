"""Shared helpers for sheet writers.

Tiny utilities only — anything doing arithmetic belongs in the waterfall /
FIFO / conversion modules. These helpers just factor out the openpyxl
boilerplate the writers would otherwise duplicate.
"""

from __future__ import annotations

from typing import Iterable, Sequence

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..design import StyleName


def write_header_row(
    ws: Worksheet,
    row: int,
    headers: Sequence[str],
    *,
    start_col: int = 1,
) -> None:
    """Write a styled header row. Uses the locked HEADER NamedStyle."""
    for offset, label in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + offset, value=label)
        cell.style = StyleName.HEADER


def apply_column_widths(ws: Worksheet, widths: Iterable[float]) -> None:
    """Set column widths in order, starting at column A.

    openpyxl's column_dimensions key expects the Excel letter ('A', 'B', …),
    so we translate via get_column_letter to avoid off-by-ones on wider sheets.
    """
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def freeze_header(ws: Worksheet, header_row: int = 1) -> None:
    """Freeze panes so the header row stays visible while scrolling."""
    ws.freeze_panes = f"A{header_row + 1}"
