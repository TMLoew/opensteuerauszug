"""Kauf/Verkauf sheet: reconstructed orders + realized gains from FIFO closes.

Two tables on one sheet: the top half shows every order (buy and sell alike)
with CHF-converted proceeds/cost and commissions; the bottom half lists the
FIFO lot closes with realized P&L. Placing them together lets a reviewer see
how each sale in the top table maps to one or more closes in the bottom.
"""

from __future__ import annotations

from openpyxl.workbook import Workbook

from ..data import TaxOverviewData
from ..design import StyleName
from ._common import apply_column_widths, freeze_header, write_header_row

SHEET_NAME = "Kauf_Verkauf"


def write_orders_sheet(workbook: Workbook, data: TaxOverviewData) -> None:
    ws = workbook.create_sheet(SHEET_NAME)
    apply_column_widths(ws, [12, 14, 22, 8, 12, 14, 14, 14, 10, 16])

    write_header_row(ws, 1, [
        "Datum", "Symbol", "ISIN", "Seite", "Menge",
        "Kurs Ø", "Bruttobetrag", "Kommission", "Währung", "Order-ID",
    ])
    freeze_header(ws, 1)

    row = 2
    for order in data.orders:
        ws.cell(row=row, column=1, value=order.earliest_fill_time.date()).style = StyleName.BODY_DATE
        ws.cell(row=row, column=2, value=order.symbol).style = StyleName.BODY_TEXT
        ws.cell(row=row, column=3, value=order.isin).style = StyleName.BODY_TEXT
        ws.cell(row=row, column=4, value=order.side).style = StyleName.BODY_TEXT
        ws.cell(row=row, column=5, value=order.total_quantity).style = StyleName.BODY_NUMBER
        ws.cell(row=row, column=6, value=order.avg_price).style = StyleName.BODY_NUMBER
        ws.cell(row=row, column=7, value=order.total_money).style = StyleName.BODY_NUMBER
        ws.cell(row=row, column=8, value=order.total_commission).style = StyleName.BODY_NUMBER
        ws.cell(row=row, column=9, value=order.currency).style = StyleName.BODY_TEXT
        ws.cell(row=row, column=10, value=order.order_id).style = StyleName.BODY_TEXT
        row += 1

    # Gap + realized-gain table header.
    row += 2
    closes_heading = ws.cell(row=row, column=1, value="Realisierte Gewinne / Verluste")
    closes_heading.style = StyleName.HEADER
    row += 1
    write_header_row(ws, row, [
        "Symbol", "ISIN", "Öffnung", "Schluss", "Menge",
        "Einstandspreis", "Verkaufspreis", "Gewinn/Verlust", "Währung", "Order-Schluss",
    ])
    row += 1
    for close in data.lot_closes:
        ws.cell(row=row, column=1, value=close.symbol).style = StyleName.BODY_TEXT
        ws.cell(row=row, column=2, value=close.isin).style = StyleName.BODY_TEXT
        ws.cell(row=row, column=3, value=close.opened_at.date()).style = StyleName.BODY_DATE
        ws.cell(row=row, column=4, value=close.closed_at.date()).style = StyleName.BODY_DATE
        ws.cell(row=row, column=5, value=close.quantity_closed).style = StyleName.BODY_NUMBER
        ws.cell(row=row, column=6, value=close.cost_per_share).style = StyleName.BODY_NUMBER
        ws.cell(row=row, column=7, value=close.proceeds_per_share).style = StyleName.BODY_NUMBER
        ws.cell(row=row, column=8, value=close.realized_pnl).style = StyleName.BODY_NUMBER
        ws.cell(row=row, column=9, value=close.currency).style = StyleName.BODY_TEXT
        ws.cell(row=row, column=10, value=close.closing_order_id).style = StyleName.BODY_TEXT
        row += 1
