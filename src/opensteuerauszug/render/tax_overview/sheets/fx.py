"""FX-Kurse sheet: every FX rate the pipeline applied, for audit.

This sheet exists so the tax clerk can reproduce any CHF figure: pick an
event date, look up its rate here, and multiply. Rates come straight from
the ESTV Kursliste in the production pipeline (phase-6 wiring).
"""

from __future__ import annotations

from openpyxl.workbook import Workbook

from ..data import TaxOverviewData
from ..design import StyleName
from ._common import apply_column_widths, freeze_header, write_header_row

SHEET_NAME = "FX_Kurse"


def write_fx_rates_sheet(workbook: Workbook, data: TaxOverviewData) -> None:
    ws = workbook.create_sheet(SHEET_NAME)
    apply_column_widths(ws, [12, 12, 16, 16])

    write_header_row(ws, 1, [
        "Datum", "Währung", "Kurs (CHF)", "Quelle",
    ])
    freeze_header(ws, 1)

    for offset, rate in enumerate(data.fx_rates, start=2):
        ws.cell(row=offset, column=1, value=rate.reference_date).style = StyleName.BODY_DATE
        ws.cell(row=offset, column=2, value=rate.currency).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=3, value=rate.rate).style = StyleName.BODY_NUMBER
        ws.cell(row=offset, column=4, value=rate.source).style = StyleName.BODY_TEXT
