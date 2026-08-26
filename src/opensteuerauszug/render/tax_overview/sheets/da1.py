"""DA-1 Hilfstabelle sheet.

The DA-1 form (Antrag auf pauschale Steueranrechnung) lets Swiss taxpayers
reclaim the portion of foreign withholding tax not already refunded at
source. This helper table is the working document that backs the DA-1
claim — one row per (security, source-country) bucket with the treaty
ceiling and recoverable amount spelled out.

This sheet displays whatever the pipeline produced; the treaty-rate lookup
and capping logic live in :class:`DA1Claim` consumers (phase-6 wiring).
"""

from __future__ import annotations

from openpyxl.workbook import Workbook

from ..data import TaxOverviewData
from ..design import StyleName
from ._common import apply_column_widths, freeze_header, write_header_row

SHEET_NAME = "DA1_Hilfstabelle"


def write_da1_sheet(workbook: Workbook, data: TaxOverviewData) -> None:
    ws = workbook.create_sheet(SHEET_NAME)
    apply_column_widths(ws, [14, 10, 32, 10, 16, 20, 16, 18, 16])

    write_header_row(
        ws,
        1,
        [
            "ISIN",
            "Symbol",
            "Bezeichnung",
            "Land",
            "Bruttoertrag (CHF)",
            "Quellensteuer (CHF)",
            "Einbehaltener Satz",
            "Abkommensobergrenze",
            "Rückforderbar (CHF)",
        ],
    )
    freeze_header(ws, 1)

    for offset, claim in enumerate(data.da1_claims, start=2):
        ws.cell(row=offset, column=1, value=claim.isin).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=2, value=claim.symbol).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=3, value=claim.description).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=4, value=claim.source_country).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=5, value=claim.gross_chf).style = StyleName.BODY_CHF
        ws.cell(row=offset, column=6, value=claim.withholding_tax_chf).style = StyleName.BODY_CHF
        ws.cell(row=offset, column=7, value=claim.withholding_rate).style = StyleName.BODY_PERCENT
        # Treaty ceiling may be unknown — leave blank rather than 0, which
        # would wrongly imply "no treaty" rather than "we didn't resolve it".
        ceiling_cell = ws.cell(row=offset, column=8, value=claim.treaty_rate_ceiling)
        ceiling_cell.style = StyleName.BODY_PERCENT
        ws.cell(row=offset, column=9, value=claim.recoverable_chf).style = StyleName.BODY_CHF
