"""Übersicht sheet: the single-page dashboard the tax clerk opens first.

Contains KPI tiles (portfolio totals) and the Vermögenszuwachs waterfall so
the reviewer can see at a glance whether the statement reconciles.
"""

from __future__ import annotations

from openpyxl.workbook import Workbook

from ..data import TaxOverviewData
from ..design import StyleName
from ._common import apply_column_widths

SHEET_NAME = "Übersicht"


def write_uebersicht_sheet(workbook: Workbook, data: TaxOverviewData) -> None:
    """Render the dashboard sheet in place.

    The sheet uses a 2-column KPI tile layout (label row + value row) so tiles
    look balanced in print preview. Waterfall lines follow beneath so the
    reviewer has one linear narrative from totals → bridge.
    """
    ws = workbook.create_sheet(SHEET_NAME, 0)  # leftmost
    apply_column_widths(ws, [32, 20, 20, 20])

    _write_title(ws, data)
    next_row = _write_kpi_tiles(ws, data, start_row=3)
    _write_waterfall(ws, data, start_row=next_row + 2)


def _write_title(ws, data: TaxOverviewData) -> None:
    title = ws.cell(
        row=1, column=1, value=f"Steuerübersicht {data.tax_year} — {data.broker.upper()}"
    )
    title.style = StyleName.KPI_VALUE  # reused: oversized + accented


def _write_kpi_tiles(ws, data: TaxOverviewData, *, start_row: int) -> int:
    """Lay out KPI tiles and return the last row used."""
    tiles = [
        ("Eröffnungswert (CHF)", data.opening_value_chf),
        ("Schlusswert (CHF)", data.closing_value_chf),
        ("Dividenden brutto (CHF)", data.total_dividends_chf()),
        ("Zinsen brutto (CHF)", data.total_interest_chf()),
        ("Quellensteuer (CHF)", data.total_withholding_tax_chf()),
        ("DA-1 rückforderbar (CHF)", data.total_da1_recoverable_chf()),
        ("Gebühren gesamt (CHF)", data.total_fees_chf()),
    ]
    # Two-column tile grid: columns 1 (label) + 2 (value).
    for offset, (label, value) in enumerate(tiles):
        row = start_row + offset
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.style = StyleName.KPI_LABEL
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.style = StyleName.KPI_VALUE
    return start_row + len(tiles) - 1


def _write_waterfall(ws, data: TaxOverviewData, *, start_row: int) -> None:
    heading = ws.cell(row=start_row, column=1, value="Vermögenszuwachs (CHF)")
    heading.style = StyleName.HEADER

    row = start_row + 1
    for line in data.waterfall.as_lines():
        label_cell = ws.cell(row=row, column=1, value=line.label)
        label_cell.style = StyleName.BODY_TEXT
        value_cell = ws.cell(row=row, column=2, value=line.amount_chf)
        # Outflows are rendered as-is (positive figures), but marked as such
        # in the label column; Swiss tax overviews do not show parentheses.
        value_cell.style = StyleName.BODY_CHF
        row += 1

    residual_label = ws.cell(row=row, column=1, value="Differenz (CHF)")
    residual_label.style = StyleName.BODY_TEXT
    residual_value = ws.cell(row=row, column=2, value=data.waterfall.residual)
    residual_value.style = StyleName.BODY_CHF
