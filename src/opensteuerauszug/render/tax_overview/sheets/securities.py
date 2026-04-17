"""Wertschriften sheet: per-ISIN closing holdings at Dec 31.

Uses the ESTV Kursliste Steuerwert per share for the CHF column where the
pipeline could resolve one. Positions without a Kursliste match fall back to
broker-reported close multiplied by the Jahresend-FX rate — this happens
upstream in the data-building step, so the sheet writer stays dumb.
"""

from __future__ import annotations

from openpyxl.workbook import Workbook

from ..data import TaxOverviewData
from ..design import StyleName
from ._common import apply_column_widths, freeze_header, write_header_row

SHEET_NAME = "Wertschriften"


def write_securities_sheet(workbook: Workbook, data: TaxOverviewData) -> None:
    ws = workbook.create_sheet(SHEET_NAME)
    apply_column_widths(ws, [14, 14, 40, 12, 10, 14, 14, 16])

    write_header_row(ws, 1, [
        "ISIN", "Symbol", "Bezeichnung", "Menge", "Währung",
        "Kurs (lokal)", "Kurs (CHF)", "Steuerwert (CHF)",
    ])
    freeze_header(ws, 1)

    for offset, position in enumerate(data.positions, start=2):
        ws.cell(row=offset, column=1, value=position.isin).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=2, value=position.symbol).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=3, value=position.description).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=4, value=position.quantity_closing).style = StyleName.BODY_NUMBER
        ws.cell(row=offset, column=5, value=position.currency).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=6, value=position.price_closing_local).style = StyleName.BODY_NUMBER
        ws.cell(row=offset, column=7, value=position.price_closing_chf).style = StyleName.BODY_NUMBER
        ws.cell(row=offset, column=8, value=position.market_value_chf).style = StyleName.BODY_CHF
