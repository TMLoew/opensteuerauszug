"""Wertschriftenverzeichnis (SG) sheet.

Direct line-by-line mapping into the Kanton SG Wertschriftenverzeichnis form:
each row is a line the taxpayer copies into PrivateTax (or the paper form).
Values are CHF, full-year aggregates.

Why a separate sheet: the Wertschriften sheet shows per-ISIN market value
only. SG wants market value *plus* income, *plus* the two withholding-tax
flavours in one table, bucketed by form field. Splitting it out keeps the
market-value sheet short and the form-oriented sheet tax-office-ready.
"""

from __future__ import annotations

from openpyxl.workbook import Workbook

from ..data import TaxOverviewData
from ..design import StyleName
from ._common import apply_column_widths, freeze_header, write_header_row

SHEET_NAME = "SG_Verzeichnis"


def write_verzeichnis_sheet(workbook: Workbook, data: TaxOverviewData) -> None:
    ws = workbook.create_sheet(SHEET_NAME)
    apply_column_widths(ws, [14, 14, 14, 36, 10, 16, 16, 18, 22])

    write_header_row(ws, 1, [
        "SG-Formular", "Typ", "ISIN", "Bezeichnung", "Menge",
        "Kurswert 31.12. (CHF)", "Ertrag brutto (CHF)",
        "Verrechnungssteuer (CHF)", "Ausl. Quellensteuer (CHF)",
    ])
    freeze_header(ws, 1)

    for offset, line in enumerate(data.verzeichnis_lines, start=2):
        ws.cell(row=offset, column=1, value=line.form_field).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=2, value=line.investment_type).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=3, value=line.isin).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=4, value=line.description).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=5, value=line.quantity).style = StyleName.BODY_NUMBER
        ws.cell(row=offset, column=6, value=line.market_value_chf).style = StyleName.BODY_CHF
        ws.cell(row=offset, column=7, value=line.income_gross_chf).style = StyleName.BODY_CHF
        ws.cell(row=offset, column=8, value=line.verrechnungssteuer_chf).style = StyleName.BODY_CHF
        ws.cell(row=offset, column=9, value=line.auslaendische_quellensteuer_chf).style = StyleName.BODY_CHF
