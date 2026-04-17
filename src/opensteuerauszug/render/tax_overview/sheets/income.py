"""Dividenden and Zinsen sheets: income event listings.

Dividends and interest share the same :class:`IncomeEvent` shape, so the
sheet writers share a single table function and differ only in the
`category` filter and the sheet label.
"""

from __future__ import annotations

from typing import Sequence

from openpyxl.workbook import Workbook

from ..data import IncomeEvent, TaxOverviewData
from ..design import StyleName
from ._common import apply_column_widths, freeze_header, write_header_row


def write_dividends_sheet(workbook: Workbook, data: TaxOverviewData) -> None:
    _write_income_sheet(workbook, "Dividenden", data.dividends)


def write_interest_sheet(workbook: Workbook, data: TaxOverviewData) -> None:
    _write_income_sheet(workbook, "Zinsen", data.interest)


def _write_income_sheet(
    workbook: Workbook, sheet_name: str, events: Sequence[IncomeEvent]
) -> None:
    ws = workbook.create_sheet(sheet_name)
    apply_column_widths(ws, [12, 14, 14, 32, 12, 10, 14, 14, 14, 14, 14])

    write_header_row(ws, 1, [
        "Zahlungsdatum", "ISIN", "Symbol", "Bezeichnung",
        "Brutto (lokal)", "Währung", "Quellensteuer (lokal)", "Netto (lokal)",
        "Brutto (CHF)", "Quellensteuer (CHF)", "Netto (CHF)",
    ])
    freeze_header(ws, 1)

    for offset, event in enumerate(events, start=2):
        ws.cell(row=offset, column=1, value=event.payment_date).style = StyleName.BODY_DATE
        ws.cell(row=offset, column=2, value=event.isin).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=3, value=event.symbol).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=4, value=event.description).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=5, value=event.gross_local).style = StyleName.BODY_NUMBER
        ws.cell(row=offset, column=6, value=event.currency).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=7, value=event.withholding_tax_local).style = StyleName.BODY_NUMBER
        ws.cell(row=offset, column=8, value=event.net_local).style = StyleName.BODY_NUMBER
        ws.cell(row=offset, column=9, value=event.gross_chf).style = StyleName.BODY_CHF
        ws.cell(row=offset, column=10, value=event.withholding_tax_chf).style = StyleName.BODY_CHF
        ws.cell(row=offset, column=11, value=event.net_chf).style = StyleName.BODY_CHF
