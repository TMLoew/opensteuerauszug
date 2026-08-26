"""Gebühren sheet: non-commission fees.

Commissions are already attributed per order on the Kauf/Verkauf sheet; this
sheet lists platform, data, wire, and other standalone fees the broker
charged during the year.
"""

from __future__ import annotations

from openpyxl.workbook import Workbook

from ..data import TaxOverviewData
from ..design import StyleName
from ._common import apply_column_widths, freeze_header, write_header_row

SHEET_NAME = "Gebühren"


def write_fees_sheet(workbook: Workbook, data: TaxOverviewData) -> None:
    ws = workbook.create_sheet(SHEET_NAME)
    apply_column_widths(ws, [12, 14, 40, 14, 10, 14])

    write_header_row(
        ws,
        1,
        [
            "Datum",
            "Art",
            "Beschreibung",
            "Betrag (lokal)",
            "Währung",
            "Betrag (CHF)",
        ],
    )
    freeze_header(ws, 1)

    for offset, fee in enumerate(data.fees, start=2):
        ws.cell(row=offset, column=1, value=fee.fee_date).style = StyleName.BODY_DATE
        ws.cell(row=offset, column=2, value=fee.kind).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=3, value=fee.description).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=4, value=fee.amount_local).style = StyleName.BODY_NUMBER
        ws.cell(row=offset, column=5, value=fee.currency).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=6, value=fee.amount_chf).style = StyleName.BODY_CHF
