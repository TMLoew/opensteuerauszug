"""Hidden KS 36 sheets (preparer-only).

Two sheets render the self-check for ESTV Kreisschreiben Nr. 36
(gewerbsmässiger Wertschriftenhandel). They are created with
``sheet_state = "hidden"`` and only appear when ``preparer_mode`` is true —
third-party exports must never ship these sheets per spec.

The traffic-light fills are applied via the KS36 NamedStyles registered in
``design.register_named_styles``; those styles are themselves marked in
``design.py`` as reserved for this file only.
"""

from __future__ import annotations

from openpyxl.workbook import Workbook

from ..data import TaxOverviewData
from ..design import StyleName
from ._common import apply_column_widths, freeze_header, write_header_row

CRITERIA_SHEET_NAME = "_KS36_Criteria"
EVIDENCE_SHEET_NAME = "_KS36_Evidence"

_STATUS_TO_STYLE = {
    "green": StyleName.KS36_GREEN,
    "amber": StyleName.KS36_AMBER,
    "red": StyleName.KS36_RED,
}


def write_ks36_criteria_sheet(workbook: Workbook, data: TaxOverviewData) -> None:
    """Write the per-criterion status table. Created hidden."""
    ws = workbook.create_sheet(CRITERIA_SHEET_NAME)
    ws.sheet_state = "hidden"
    apply_column_widths(ws, [40, 14, 14, 12, 12, 40])

    write_header_row(ws, 1, [
        "Kriterium", "Beobachtet", "Schwelle", "Einheit", "Status", "Hinweis",
    ])
    freeze_header(ws, 1)

    for offset, criterion in enumerate(data.ks36_criteria, start=2):
        ws.cell(row=offset, column=1, value=criterion.label).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=2, value=criterion.observed_value).style = StyleName.BODY_NUMBER
        ws.cell(row=offset, column=3, value=criterion.threshold).style = StyleName.BODY_NUMBER
        ws.cell(row=offset, column=4, value=criterion.unit).style = StyleName.BODY_TEXT
        status_cell = ws.cell(row=offset, column=5, value=criterion.status)
        status_cell.style = _STATUS_TO_STYLE.get(
            criterion.status, StyleName.BODY_TEXT
        )
        ws.cell(row=offset, column=6, value=criterion.note or "").style = StyleName.BODY_TEXT


def write_ks36_evidence_sheet(workbook: Workbook, data: TaxOverviewData) -> None:
    """Write supporting evidence rows. Created hidden."""
    ws = workbook.create_sheet(EVIDENCE_SHEET_NAME)
    ws.sheet_state = "hidden"
    apply_column_widths(ws, [20, 20, 40, 14, 12])

    write_header_row(ws, 1, [
        "Kriterium", "Kategorie", "Beschreibung", "Betrag (CHF)", "Datum",
    ])
    freeze_header(ws, 1)

    for offset, evidence in enumerate(data.ks36_evidence, start=2):
        ws.cell(row=offset, column=1, value=evidence.criterion_code).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=2, value=evidence.category).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=3, value=evidence.description).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=4, value=evidence.value_chf).style = StyleName.BODY_CHF
        ws.cell(row=offset, column=5, value=evidence.evidence_date).style = StyleName.BODY_DATE
