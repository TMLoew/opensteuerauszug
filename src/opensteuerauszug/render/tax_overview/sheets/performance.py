"""Performance sheet: per-position P&L, sector / currency splits, benchmarks.

Consumes the pre-computed :class:`PerformanceSection` on
:class:`TaxOverviewData`; all arithmetic lives in the pipeline.  When no
performance payload is attached (e.g. preparer tests that stub the
pipeline) the sheet is skipped entirely rather than emitting an empty tab,
so downstream consumers don't see a placeholder.

Charts are native openpyxl chart objects (BarChart, PieChart) anchored to
cells beside the tables. They reference the same table ranges the reviewer
sees, so edits to the underlying numbers flow through to the charts when
the workbook is reopened in Excel / LibreOffice.
"""

from __future__ import annotations

from decimal import Decimal
from typing import List

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.workbook import Workbook

from ..data import TaxOverviewData
from ..design import PALETTE, StyleName
from ._common import apply_column_widths, freeze_header, write_header_row


SHEET_NAME = "Performance"

ZERO = Decimal("0")
HUNDRED = Decimal("100")


def write_performance_sheet(workbook: Workbook, data: TaxOverviewData) -> None:
    perf = data.performance
    if perf is None:
        return

    ws = workbook.create_sheet(SHEET_NAME)
    apply_column_widths(
        ws,
        [14, 14, 40, 20, 10, 14, 14, 14, 14, 14, 14, 14, 14, 10],
    )

    row = _write_summary(ws, perf, start_row=1)
    row += 1
    positions_range = _write_positions_table(ws, perf, start_row=row)
    row = positions_range["end"] + 2
    sectors_range = _write_allocation_table(
        ws, perf.sectors, title="Sektor-Aggregation", start_row=row,
    )
    row = sectors_range["end"] + 2
    currencies_range = _write_allocation_table(
        ws, perf.currencies, title="Währungs-Aggregation", start_row=row,
    )
    row = currencies_range["end"] + 2
    benchmark_range = None
    if perf.benchmarks:
        benchmark_range = _write_benchmark_table_with_portfolio(
            ws, perf, start_row=row,
        )
        row = benchmark_range["end"] + 2

    _add_charts(
        ws, perf,
        positions_range=positions_range,
        sectors_range=sectors_range,
        currencies_range=currencies_range,
        benchmark_range=benchmark_range,
    )


def _write_summary(ws, perf, *, start_row: int) -> int:
    summary = perf.summary
    cash_label = "Schluss Cash (CHF)"
    if not summary.cash_known:
        cash_label += " — Eröffnung n/v"
    tiles = [
        ("Eröffnung Wertschriften (CHF)", summary.opening_value_chf),
        ("Schluss Wertschriften (CHF)", summary.closing_value_chf),
        (cash_label, summary.closing_cash_chf),
        ("Einzahlungen (CHF)", summary.deposits_gross_chf),
        ("Auszahlungen (CHF)", summary.withdrawals_chf),
        ("Netto-Einzahlungen (CHF)", summary.net_deposits_chf),
        ("Gesamt-P&L Wertschriften (CHF)", summary.total_pnl_chf),
        ("Dividenden (CHF)", summary.dividends_chf),
        ("Zinsen (CHF)", summary.interest_chf),
        ("Gebühren (CHF)", summary.fees_chf),
        ("Rendite Modified Dietz (%)", _pct_or_none(summary.money_weighted_return_pct)),
        ("Rendite einfach (%)", _pct_or_none(summary.simple_return_pct)),
    ]
    for offset, (label, value) in enumerate(tiles):
        row = start_row + offset
        ws.cell(row=row, column=1, value=label).style = StyleName.KPI_LABEL
        value_cell = ws.cell(row=row, column=2, value=value)
        if isinstance(value, Decimal) and label.endswith("(%)"):
            value_cell.style = StyleName.BODY_PERCENT
        else:
            value_cell.style = StyleName.KPI_VALUE
    return start_row + len(tiles) - 1


def _write_positions_table(ws, perf, *, start_row: int) -> dict:
    """Return the table's row range as ``{"start", "header", "end"}``.

    ``start`` is the heading row, ``header`` the column-header row, ``end``
    the last data row (so ranges for charts use ``header+1`` .. ``end``).
    """
    heading = ws.cell(row=start_row, column=1, value="Positionen")
    heading.style = StyleName.HEADER
    header_row = start_row + 1
    write_header_row(ws, header_row, [
        "ISIN", "Symbol", "Bezeichnung", "Sektor", "Währung",
        "Eröffnung CHF", "Schluss CHF", "Käufe CHF", "Verkäufe CHF",
        "Dividenden CHF", "Realisiert CHF", "Unrealisiert CHF",
        "Gesamt-P&L CHF", "Rendite %",
    ])
    freeze_header(ws, header_row)
    for offset, p in enumerate(perf.positions, start=header_row + 1):
        ws.cell(row=offset, column=1, value=p.isin or "").style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=2, value=p.symbol).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=3, value=p.description).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=4, value=p.sector).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=5, value=p.currency).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=6, value=p.opening_value_chf).style = StyleName.BODY_CHF
        ws.cell(row=offset, column=7, value=p.closing_value_chf).style = StyleName.BODY_CHF
        ws.cell(row=offset, column=8, value=p.buys_chf).style = StyleName.BODY_CHF
        ws.cell(row=offset, column=9, value=p.sells_chf).style = StyleName.BODY_CHF
        ws.cell(row=offset, column=10, value=p.dividends_chf).style = StyleName.BODY_CHF
        ws.cell(row=offset, column=11, value=p.realized_pnl_chf).style = StyleName.BODY_CHF
        ws.cell(row=offset, column=12, value=p.unrealized_pnl_chf).style = StyleName.BODY_CHF
        ws.cell(row=offset, column=13, value=p.total_pnl_chf).style = StyleName.BODY_CHF
        ws.cell(row=offset, column=14, value=_pct_or_none(p.return_pct)).style = StyleName.BODY_PERCENT
    end_row = header_row + len(perf.positions)
    return {"start": start_row, "header": header_row, "end": end_row}


def _write_allocation_table(ws, allocations, *, title: str, start_row: int) -> dict:
    heading = ws.cell(row=start_row, column=1, value=title)
    heading.style = StyleName.HEADER
    header_row = start_row + 1
    write_header_row(
        ws, header_row,
        ["Bezeichnung", "Marktwert CHF", "Gewicht %", "P&L CHF"],
    )
    for offset, a in enumerate(allocations, start=header_row + 1):
        ws.cell(row=offset, column=1, value=a.label).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=2, value=a.market_value_chf).style = StyleName.BODY_CHF
        ws.cell(row=offset, column=3, value=a.weight_pct / HUNDRED).style = StyleName.BODY_PERCENT
        ws.cell(row=offset, column=4, value=a.pnl_chf).style = StyleName.BODY_CHF
    end_row = header_row + len(allocations)
    return {"start": start_row, "header": header_row, "end": end_row}


def _write_benchmark_table_with_portfolio(ws, perf, *, start_row: int) -> dict:
    """Benchmarks table prepended with the portfolio return.

    Ordering matters for the chart: the portfolio sits at the top as the
    protagonist; benchmarks follow in catalog order. Returns the usual
    start/header/end row dict so the chart can reference header+1..end.
    """
    heading = ws.cell(row=start_row, column=1, value="Benchmark-Referenzen")
    heading.style = StyleName.HEADER
    header_row = start_row + 1
    write_header_row(ws, header_row, ["Kürzel", "Bezeichnung", "Rendite %", "Hinweis"])

    portfolio_return = perf.summary.money_weighted_return_pct
    rows: List[tuple[str, str, Decimal | None, str]] = []
    if portfolio_return is not None:
        rows.append((
            "PORTFOLIO",
            "Portfolio (Modified Dietz)",
            portfolio_return,
            "Errechnet aus Eröffnung, Schluss und Netto-Einzahlungen.",
        ))
    for b in perf.benchmarks:
        rows.append((b.code, b.label, b.return_pct, b.note or ""))

    for offset, (code, label, ret, note) in enumerate(rows, start=header_row + 1):
        ws.cell(row=offset, column=1, value=code).style = StyleName.BODY_TEXT
        ws.cell(row=offset, column=2, value=label).style = StyleName.BODY_TEXT
        value = _pct_or_none(ret) if ret is not None else None
        ws.cell(row=offset, column=3, value=value).style = StyleName.BODY_PERCENT
        ws.cell(row=offset, column=4, value=note).style = StyleName.BODY_TEXT

    end_row = header_row + len(rows)
    return {"start": start_row, "header": header_row, "end": end_row}


# ---------------------------------------------------------------------------
# Charts
# ---------------------------------------------------------------------------


def _add_charts(
    ws, perf, *,
    positions_range, sectors_range, currencies_range, benchmark_range,
) -> None:
    """Attach the four chart visuals beside the tables.

    Charts are anchored in column P (well right of the 14-column table
    area) so they don't cover the data; the reviewer can scroll horizontally
    to see them, or shrink the window to force-tile. Excel treats each
    chart as a floating object — they move with the cell reference they're
    anchored to.
    """
    anchor_col = "P"
    row = 1

    if benchmark_range is not None and benchmark_range["end"] > benchmark_range["header"]:
        chart = _build_benchmark_chart(ws, benchmark_range)
        ws.add_chart(chart, f"{anchor_col}{row}")
        row += 20

    if sectors_range["end"] > sectors_range["header"]:
        chart = _build_pie_chart(
            ws,
            title="Sektor-Allokation",
            range_info=sectors_range,
        )
        ws.add_chart(chart, f"{anchor_col}{row}")
        row += 20

    if currencies_range["end"] > currencies_range["header"]:
        chart = _build_pie_chart(
            ws,
            title="Währungs-Allokation",
            range_info=currencies_range,
        )
        ws.add_chart(chart, f"{anchor_col}{row}")
        row += 20

    if positions_range["end"] > positions_range["header"]:
        chart = _build_top_positions_chart(ws, perf, positions_range)
        if chart is not None:
            ws.add_chart(chart, f"{anchor_col}{row}")


def _build_benchmark_chart(ws, range_info: dict) -> BarChart:
    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = "Portfolio vs. Benchmarks"
    chart.y_axis.title = None
    chart.x_axis.title = "Rendite"
    chart.legend = None
    chart.height = 9
    chart.width = 18

    header = range_info["header"]
    end = range_info["end"]
    # Column 3 holds Rendite %, column 2 holds the label.
    values = Reference(
        ws, min_col=3, max_col=3, min_row=header, max_row=end,
    )
    categories = Reference(
        ws, min_col=2, max_col=2, min_row=header + 1, max_row=end,
    )
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    chart.dataLabels = DataLabelList(showVal=True)
    return chart


def _build_pie_chart(ws, *, title: str, range_info: dict) -> PieChart:
    chart = PieChart()
    chart.title = title
    chart.height = 9
    chart.width = 14

    header = range_info["header"]
    end = range_info["end"]
    labels = Reference(
        ws, min_col=1, max_col=1, min_row=header + 1, max_row=end,
    )
    # Market value drives the slice extents.
    values = Reference(
        ws, min_col=2, max_col=2, min_row=header, max_row=end,
    )
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(labels)
    chart.dataLabels = DataLabelList(showPercent=True)
    return chart


def _build_top_positions_chart(ws, perf, positions_range: dict):
    """Bar chart of the top P&L contributions.

    Uses columns 2 (Symbol) as category and 13 (Gesamt-P&L CHF) as value.
    Excel's horizontal bar chart auto-handles signed values, so negative
    contributions extend to the left. The column range mirrors what's in
    the sheet — no synthesis needed.
    """
    if not perf.positions:
        return None

    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = "Top-Beiträge zum Ergebnis (CHF)"
    chart.legend = None
    chart.height = 14
    chart.width = 22

    header = positions_range["header"]
    # Sheet writes positions already sorted by abs(P&L) DESC (build_perf_section),
    # so the first ~12 rows are the top contributors.
    end = min(positions_range["end"], header + 12)

    values = Reference(
        ws, min_col=13, max_col=13, min_row=header, max_row=end,
    )
    categories = Reference(
        ws, min_col=2, max_col=2, min_row=header + 1, max_row=end,
    )
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    chart.dataLabels = DataLabelList(showVal=True)
    return chart


def _pct_or_none(value):
    """BODY_PERCENT style multiplies by 100; convert dashboard-percent to rate."""
    if value is None:
        return None
    return value / HUNDRED
