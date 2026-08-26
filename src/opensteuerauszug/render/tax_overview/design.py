"""Design tokens for the tax-overview mode.

Single source of truth for colours, typography, number formats, and reusable
openpyxl styles. Every workbook, HTML, and PDF writer imports from here so the
locked palette is enforced in one place.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List

from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side

# ---------------------------------------------------------------------------
# Locked palette (see spec "Visual design system -> Colour palette")
# ---------------------------------------------------------------------------

PALETTE: Dict[str, str] = {
    "ink": "#0F1720",
    "ink_muted": "#4A5568",
    "paper": "#FFFFFF",
    "paper_warm": "#FAF8F3",
    "rule": "#E2E4E8",
    "primary": "#1B3A5B",
    "primary_80": "#3C5F84",
    "accent": "#A47551",
    "positive": "#2E6F4E",
    "negative": "#8A2A2A",
    "warn": "#B08900",
}

# Categorical series colours for charts (pie slices, bars), in fixed assignment
# order — never cycled or re-sorted. Distinct from the status colours above
# (positive/negative/warn stay reserved for state). The order and values pass
# a colour-vision-deficiency / lightness / chroma / contrast validation on a
# white surface; if you change a value or the order, re-validate.
CHART_SERIES: List[str] = [
    "#2E6BA8",  # blue
    "#C97F3C",  # ochre
    "#6B5CA8",  # violet
    "#2F8F6B",  # green
    "#B58E0A",  # gold
    "#A5504B",  # muted red
]


# openpyxl wants RGB without the leading '#'.
def _rgb(token: str) -> str:
    return PALETTE[token].lstrip("#")


# ---------------------------------------------------------------------------
# Swiss number & date formats (see spec "Swiss number & date formatting")
# ---------------------------------------------------------------------------

# Swiss locale uses apostrophe as thousands separator and dot as decimal.
# Excel format codes only support "," as the (repeating) group separator; the
# separator glyph itself comes from the viewer's locale, so de-CH Excel renders
# these with apostrophes. (A literal "'" in the code would group only once.)
NUMBER_FORMAT_CHF = '"CHF "#,##0.00;-"CHF "#,##0.00;"CHF "0.00'
NUMBER_FORMAT_CURRENCY_GENERIC = '@" "#,##0.00'  # prefix column provides ISO code
NUMBER_FORMAT_PLAIN = "#,##0.00;-#,##0.00;0.00"
NUMBER_FORMAT_INTEGER = "#,##0;-#,##0;0"
NUMBER_FORMAT_PERCENT = '0.0%;-0.0%;0.0%'
NUMBER_FORMAT_DATE = 'DD.MM.YYYY'
NUMBER_FORMAT_DATETIME = 'DD.MM.YYYY HH:MM'


# ---------------------------------------------------------------------------
# Typography (Excel uses Calibri; HTML/PDF apply Inter/Source Serif separately)
# ---------------------------------------------------------------------------

FONT_FAMILY_EXCEL = "Calibri"

FONT_BODY = Font(name=FONT_FAMILY_EXCEL, size=11, color=_rgb("ink"))
FONT_BODY_MUTED = Font(name=FONT_FAMILY_EXCEL, size=11, color=_rgb("ink_muted"))
FONT_HEADER = Font(name=FONT_FAMILY_EXCEL, size=11, bold=True, color="FFFFFF")
FONT_SECTION_TITLE = Font(name=FONT_FAMILY_EXCEL, size=14, bold=True, color=_rgb("ink"))
FONT_KPI_LABEL = Font(name=FONT_FAMILY_EXCEL, size=9, bold=False, color=_rgb("ink_muted"))
FONT_KPI_VALUE = Font(name=FONT_FAMILY_EXCEL, size=22, bold=True, color=_rgb("ink"))
FONT_POSITIVE = Font(name=FONT_FAMILY_EXCEL, size=11, color=_rgb("positive"))
FONT_NEGATIVE = Font(name=FONT_FAMILY_EXCEL, size=11, color=_rgb("negative"))

# Ampel colours are reserved for hidden KS36 sheets only (spec: visible workbook
# never uses traffic-light fills). Exposed here as distinct constants so it's
# obvious where they're allowed.
FILL_AMPEL_GREEN = PatternFill("solid", fgColor="C6E0B4")
FILL_AMPEL_AMBER = PatternFill("solid", fgColor="FFE699")
FILL_AMPEL_RED = PatternFill("solid", fgColor="F4B3B3")

FILL_HEADER = PatternFill("solid", fgColor=_rgb("primary"))
FILL_ZEBRA = PatternFill("solid", fgColor=_rgb("paper_warm"))
FILL_KPI_TILE = PatternFill("solid", fgColor=_rgb("paper_warm"))


# ---------------------------------------------------------------------------
# Borders
# ---------------------------------------------------------------------------

_rule_side = Side(style="thin", color=_rgb("rule"))
BORDER_ROW_BOTTOM = Border(bottom=_rule_side)


# ---------------------------------------------------------------------------
# Alignments
# ---------------------------------------------------------------------------

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_HEADER = Alignment(horizontal="left", vertical="center", wrap_text=False)


# ---------------------------------------------------------------------------
# NamedStyle builders
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class StyleName:
    """String constants for NamedStyle lookups — keeps sheet-writer code readable."""

    HEADER = "tax_overview_header"
    BODY_TEXT = "tax_overview_body_text"
    BODY_CHF = "tax_overview_body_chf"
    BODY_NUMBER = "tax_overview_body_number"
    BODY_INTEGER = "tax_overview_body_integer"
    BODY_PERCENT = "tax_overview_body_percent"
    BODY_DATE = "tax_overview_body_date"
    KPI_LABEL = "tax_overview_kpi_label"
    KPI_VALUE = "tax_overview_kpi_value"
    # Traffic-light styles — reserved for hidden KS36 sheets only. Using these
    # on a visible sheet violates the spec's third-party-safety rule.
    KS36_GREEN = "tax_overview_ks36_green"
    KS36_AMBER = "tax_overview_ks36_amber"
    KS36_RED = "tax_overview_ks36_red"


def _named(
    name: str,
    *,
    font: Font,
    fill: PatternFill | None = None,
    number_format: str | None = None,
    alignment: Alignment = ALIGN_LEFT,
    border: Border | None = BORDER_ROW_BOTTOM,
) -> NamedStyle:
    style = NamedStyle(name=name)
    style.font = font
    style.alignment = alignment
    if fill is not None:
        style.fill = fill
    if number_format is not None:
        style.number_format = number_format
    if border is not None:
        style.border = border
    return style


def register_named_styles(workbook) -> None:
    """Register all tax-overview NamedStyles on the workbook, idempotent.

    Call once per workbook before writing cells. Cells then reference the style
    by name: ``cell.style = StyleName.BODY_CHF``.
    """
    # openpyxl exposes named_styles as a list of names (strings), not objects.
    existing = set(workbook.named_styles)
    defs = [
        _named(
            StyleName.HEADER,
            font=FONT_HEADER,
            fill=FILL_HEADER,
            alignment=ALIGN_HEADER,
            border=None,
        ),
        _named(StyleName.BODY_TEXT, font=FONT_BODY, alignment=ALIGN_LEFT),
        _named(
            StyleName.BODY_CHF,
            font=FONT_BODY,
            number_format=NUMBER_FORMAT_CHF,
            alignment=ALIGN_RIGHT,
        ),
        _named(
            StyleName.BODY_NUMBER,
            font=FONT_BODY,
            number_format=NUMBER_FORMAT_PLAIN,
            alignment=ALIGN_RIGHT,
        ),
        _named(
            StyleName.BODY_INTEGER,
            font=FONT_BODY,
            number_format=NUMBER_FORMAT_INTEGER,
            alignment=ALIGN_RIGHT,
        ),
        _named(
            StyleName.BODY_PERCENT,
            font=FONT_BODY,
            number_format=NUMBER_FORMAT_PERCENT,
            alignment=ALIGN_RIGHT,
        ),
        _named(
            StyleName.BODY_DATE,
            font=FONT_BODY,
            number_format=NUMBER_FORMAT_DATE,
            alignment=ALIGN_CENTER,
        ),
        _named(
            StyleName.KPI_LABEL,
            font=FONT_KPI_LABEL,
            fill=FILL_KPI_TILE,
            alignment=ALIGN_LEFT,
            border=None,
        ),
        _named(
            StyleName.KPI_VALUE,
            font=FONT_KPI_VALUE,
            fill=FILL_KPI_TILE,
            number_format=NUMBER_FORMAT_CHF,
            alignment=ALIGN_RIGHT,
            border=None,
        ),
        _named(
            StyleName.KS36_GREEN,
            font=FONT_BODY,
            fill=FILL_AMPEL_GREEN,
            alignment=ALIGN_LEFT,
            border=None,
        ),
        _named(
            StyleName.KS36_AMBER,
            font=FONT_BODY,
            fill=FILL_AMPEL_AMBER,
            alignment=ALIGN_LEFT,
            border=None,
        ),
        _named(
            StyleName.KS36_RED,
            font=FONT_BODY,
            fill=FILL_AMPEL_RED,
            alignment=ALIGN_LEFT,
            border=None,
        ),
    ]
    for style in defs:
        if style.name not in existing:
            workbook.add_named_style(style)


# ---------------------------------------------------------------------------
# CSS variables (consumed by the HTML writer)
# ---------------------------------------------------------------------------


def css_variables() -> str:
    """Return a CSS :root block with the locked palette + typography tokens.

    Emitted verbatim into the HTML writer's <style> block so the HTML report
    and the workbook share one source of truth.
    """
    lines = [":root {"]
    for key, hex_value in PALETTE.items():
        lines.append(f"    --color-{key.replace('_', '-')}: {hex_value};")
    lines.extend(
        [
            "    --font-body: 'Inter', 'Helvetica Neue', system-ui, sans-serif;",
            "    --font-display: 'Source Serif 4', 'Georgia', serif;",
            "    --font-num: 'Inter', 'Helvetica Neue', system-ui, sans-serif;",
            "    --grid-unit: 8px;",
            "}",
        ]
    )
    return "\n".join(lines)
