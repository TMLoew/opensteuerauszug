"""Phase-1 smoke tests: CLI wiring, design tokens, and stub workbook writer.

Deeper tests (order reconstruction, FIFO, KS36 gating) land in their own phases.
These tests only guarantee the scaffold exists and produces valid files.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from typer.testing import CliRunner

from opensteuerauszug.render.tax_overview import (
    ALL_FORMATS,
    TaxOverviewRequest,
    compute_report_hash,
    parse_formats,
    write_tax_overview,
)
from opensteuerauszug.render.tax_overview.data import TaxOverviewData
from opensteuerauszug.render.tax_overview.design import (
    PALETTE,
    StyleName,
    css_variables,
    register_named_styles,
)
from opensteuerauszug.render.tax_overview.pipeline import PipelineResult
from opensteuerauszug.render.tax_overview.waterfall import Waterfall
from opensteuerauszug.steuerauszug import app


@pytest.fixture
def dummy_statement(tmp_path: Path) -> Path:
    path = tmp_path / "stub_statement.xml"
    path.write_text("<FlexQueryResponse/>\n", encoding="utf-8")
    return path


def _empty_tax_overview_data(*, preparer_mode: bool, tax_year: int) -> TaxOverviewData:
    return TaxOverviewData(
        tax_year=tax_year,
        broker="ibkr",
        preparer_mode=preparer_mode,
        opening_value_chf=Decimal("0"),
        closing_value_chf=Decimal("0"),
        waterfall=Waterfall(
            opening=Decimal("0"), inflows=[], outflows=[], closing=Decimal("0")
        ),
        positions=[], orders=[], lot_closes=[],
        dividends=[], interest=[], fees=[], fx_rates=[],
        verzeichnis_lines=[], da1_claims=[],
        ks36_criteria=[], ks36_evidence=[],
    )


@pytest.fixture(autouse=True)
def _stub_pipeline(monkeypatch: pytest.MonkeyPatch) -> None:
    """Bypass the real broker+Kursliste pipeline in scaffold tests.

    Phase-1 tests exercise the writer wiring (format dispatch, KS36 gating,
    CLI surface) — not pipeline correctness. Feeding them a minimal
    TaxOverviewData keeps the tests hermetic and fast.
    """

    def _fake_build(*, input_path: Path, broker: str, tax_year: int,
                    preparer_mode: bool = False, **_kwargs) -> PipelineResult:
        from opensteuerauszug.model.ech0196 import TaxStatement
        data = _empty_tax_overview_data(preparer_mode=preparer_mode, tax_year=tax_year)
        # The writer consumes only `data`; skip validation so we don't have to
        # synthesise an entire eCH-0196 tree just to satisfy pydantic.
        return PipelineResult(data=data, statement=TaxStatement.model_construct())

    monkeypatch.setattr(
        "opensteuerauszug.render.tax_overview.writer.build_tax_overview_data",
        _fake_build,
    )


@pytest.fixture
def request_all_formats(dummy_statement: Path, tmp_path: Path) -> TaxOverviewRequest:
    return TaxOverviewRequest(
        input_path=dummy_statement,
        broker="ibkr",
        tax_year=2025,
        output_dir=tmp_path / "out",
        formats=ALL_FORMATS,
        preparer_mode=False,
    )


def test_palette_is_locked_set_of_tokens() -> None:
    expected = {
        "ink", "ink_muted", "paper", "paper_warm", "rule",
        "primary", "primary_80", "accent",
        "positive", "negative", "warn",
    }
    assert set(PALETTE.keys()) == expected
    for name, value in PALETTE.items():
        assert value.startswith("#") and len(value) == 7, f"{name} must be #RRGGBB"


def test_css_variables_exposes_every_palette_token() -> None:
    css = css_variables()
    assert css.startswith(":root {")
    for token, hex_value in PALETTE.items():
        css_name = token.replace("_", "-")
        assert f"--color-{css_name}: {hex_value};" in css


def test_register_named_styles_is_idempotent() -> None:
    from openpyxl import Workbook
    wb = Workbook()
    register_named_styles(wb)
    first = set(wb.named_styles)
    register_named_styles(wb)  # should not raise / duplicate
    assert set(wb.named_styles) == first
    for name in (StyleName.HEADER, StyleName.BODY_CHF, StyleName.KPI_VALUE):
        assert name in first


def test_parse_formats_defaults_to_all_when_none() -> None:
    assert parse_formats(None) == ALL_FORMATS
    assert parse_formats("") == ALL_FORMATS


def test_parse_formats_preserves_canonical_order() -> None:
    assert parse_formats("pdf,xlsx") == ("xlsx", "pdf")
    assert parse_formats("html") == ("html",)


def test_parse_formats_rejects_unknown() -> None:
    with pytest.raises(ValueError, match="unknown output format"):
        parse_formats("xlsx,csv")


def test_compute_report_hash_is_stable_across_modes(dummy_statement: Path, tmp_path: Path) -> None:
    base_kwargs = dict(
        input_path=dummy_statement,
        broker="ibkr",
        tax_year=2025,
        output_dir=tmp_path,
        formats=ALL_FORMATS,
    )
    h1 = compute_report_hash(TaxOverviewRequest(**base_kwargs, preparer_mode=False))
    h2 = compute_report_hash(TaxOverviewRequest(**base_kwargs, preparer_mode=True))
    assert h1 == h2, "hash must not depend on preparer_mode (spec default #3)"


def test_compute_report_hash_changes_with_input(tmp_path: Path) -> None:
    a = tmp_path / "a.xml"
    b = tmp_path / "b.xml"
    a.write_text("<A/>", encoding="utf-8")
    b.write_text("<B/>", encoding="utf-8")
    common = dict(broker="ibkr", tax_year=2025, output_dir=tmp_path,
                  formats=ALL_FORMATS, preparer_mode=False)
    ha = compute_report_hash(TaxOverviewRequest(input_path=a, **common))
    hb = compute_report_hash(TaxOverviewRequest(input_path=b, **common))
    assert ha != hb


def test_write_tax_overview_produces_requested_files(request_all_formats: TaxOverviewRequest) -> None:
    produced = write_tax_overview(request_all_formats)
    assert len(produced) == 3
    suffixes = sorted(p.suffix for p in produced)
    assert suffixes == [".html", ".pdf", ".xlsx"]
    for path in produced:
        assert path.exists()
        assert path.stat().st_size > 0


def test_xlsx_output_has_uebersicht_sheet_and_no_hidden_ks36(
    request_all_formats: TaxOverviewRequest,
) -> None:
    produced = write_tax_overview(request_all_formats)
    xlsx = next(p for p in produced if p.suffix == ".xlsx")
    wb = load_workbook(xlsx)
    assert "Übersicht" in wb.sheetnames
    # Preparer mode is off, so no KS36 sheet should exist at all.
    assert not any(name.startswith("_KS36") for name in wb.sheetnames)


def test_cli_help_lists_tax_overview_subcommand() -> None:
    runner = CliRunner()
    result = runner.invoke(app, ["--help"])
    assert result.exit_code == 0
    assert "tax-overview" in result.stdout


def test_cli_tax_overview_help_shows_flags() -> None:
    runner = CliRunner()
    result = runner.invoke(app, ["tax-overview", "--help"])
    assert result.exit_code == 0
    for flag in ("--input", "--broker", "--year", "--output-dir", "--formats", "--preparer-mode"):
        assert flag in result.stdout


def test_cli_tax_overview_rejects_unknown_broker(dummy_statement: Path, tmp_path: Path) -> None:
    runner = CliRunner()
    result = runner.invoke(
        app,
        [
            "tax-overview",
            "--input", str(dummy_statement),
            "--broker", "fidelity",
            "--year", "2025",
            "--output-dir", str(tmp_path / "out"),
        ],
    )
    assert result.exit_code != 0
    assert "fidelity" in result.stdout or "fidelity" in (result.stderr or "")


def test_cli_tax_overview_end_to_end(dummy_statement: Path, tmp_path: Path) -> None:
    out_dir = tmp_path / "out"
    runner = CliRunner()
    result = runner.invoke(
        app,
        [
            "tax-overview",
            "--input", str(dummy_statement),
            "--broker", "ibkr",
            "--year", "2025",
            "--output-dir", str(out_dir),
            "--formats", "xlsx",
        ],
    )
    assert result.exit_code == 0, result.stdout
    assert (out_dir / "tax_overview_2025.xlsx").exists()
