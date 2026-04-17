"""Phase-5 tests: visible workbook sheet writers.

Verifies sheet presence, order, header rows, data values, and NamedStyle
application. The data fixture is hand-built so tests remain stable even as
the upstream (importer → TaxOverviewData) pipeline evolves in phase 6+.
"""

from __future__ import annotations

from datetime import date, datetime, timedelta
from decimal import Decimal

import pytest
from openpyxl import Workbook

from opensteuerauszug.render.tax_overview.data import (
    FeeEvent,
    FXRateUsed,
    IncomeEvent,
    PositionSummary,
    TaxOverviewData,
)
from opensteuerauszug.render.tax_overview.design import StyleName
from opensteuerauszug.render.tax_overview.fifo import LotClose
from opensteuerauszug.render.tax_overview.orders import Fill, reconstruct_orders
from opensteuerauszug.render.tax_overview.render import render_workbook
from opensteuerauszug.render.tax_overview.waterfall import (
    WaterfallLine,
    build_waterfall,
)


D = Decimal
T0 = datetime(2025, 3, 14, 10, 0, 0)


# ---------------------------------------------------------------------------
# Fixture: a small but realistic TaxOverviewData
# ---------------------------------------------------------------------------


@pytest.fixture
def sample_data() -> TaxOverviewData:
    buy_fill = Fill(
        fill_id="f1", symbol="AAPL", side="BUY", quantity=D("10"),
        price=D("180"), money=D("1800"), commission=D("-1"),
        currency="USD", trade_time=T0, asset_category="STK",
        isin="US0378331005", conid="265598", ib_order_id="IB-1",
    )
    sell_fill = Fill(
        fill_id="f2", symbol="AAPL", side="SELL", quantity=D("-10"),
        price=D("210"), money=D("2100"), commission=D("-1"),
        currency="USD", trade_time=T0 + timedelta(days=90),
        asset_category="STK", isin="US0378331005", conid="265598",
        ib_order_id="IB-2",
    )
    orders = reconstruct_orders([buy_fill, sell_fill])

    close = LotClose(
        lot_id="lot:ib:IB-1",
        symbol="AAPL",
        isin="US0378331005",
        quantity_closed=D("10"),
        cost_per_share=D("180.1"),
        proceeds_per_share=D("209.9"),
        currency="USD",
        opened_at=T0,
        closed_at=T0 + timedelta(days=90),
        opening_order_id="ib:IB-1",
        closing_order_id="ib:IB-2",
    )

    dividend = IncomeEvent(
        payment_date=date(2025, 5, 15),
        isin="US0378331005", symbol="AAPL",
        description="Apple Inc. Q1 dividend",
        category="dividend",
        gross_local=D("25.00"), currency="USD",
        withholding_tax_local=D("3.75"),
        net_local=D("21.25"),
        gross_chf=D("22.50"),
        withholding_tax_chf=D("3.38"),
        net_chf=D("19.12"),
    )
    interest = IncomeEvent(
        payment_date=date(2025, 7, 1),
        isin=None, symbol="CASH-USD",
        description="IBKR credit interest",
        category="interest",
        gross_local=D("12.00"), currency="USD",
        withholding_tax_local=D("0"),
        net_local=D("12.00"),
        gross_chf=D("10.80"),
        withholding_tax_chf=D("0"),
        net_chf=D("10.80"),
    )
    fee = FeeEvent(
        fee_date=date(2025, 12, 31),
        kind="platform",
        description="Market data subscription",
        amount_local=D("10"), currency="USD",
        amount_chf=D("9.00"),
    )
    fx = FXRateUsed(
        currency="USD",
        reference_date=date(2025, 5, 15),
        rate=D("0.90"),
    )
    position = PositionSummary(
        isin="US0378331005", symbol="AAPL",
        description="Apple Inc.",
        quantity_closing=D("0"),  # everything sold in this fixture
        currency="USD",
        price_closing_local=D("210"),
        price_closing_chf=D("189"),
        market_value_chf=D("0"),
    )

    waterfall = build_waterfall(
        opening=D("100000"),
        closing=D("115000"),
        inflows=[
            WaterfallLine("Dividends", D("22.50"), "inflow"),
            WaterfallLine("Interest", D("10.80"), "inflow"),
            WaterfallLine("Realized + unrealized", D("14976.70"), "inflow"),
        ],
        outflows=[
            WaterfallLine("Withholding tax", D("3.38"), "outflow"),
            WaterfallLine("Fees (incl. commissions)", D("6.62"), "outflow"),
        ],
    )

    return TaxOverviewData(
        tax_year=2025,
        broker="ibkr",
        preparer_mode=False,
        opening_value_chf=D("100000"),
        closing_value_chf=D("115000"),
        waterfall=waterfall,
        positions=[position],
        orders=orders,
        lot_closes=[close],
        dividends=[dividend],
        interest=[interest],
        fees=[fee],
        fx_rates=[fx],
    )


# ---------------------------------------------------------------------------
# Workbook-level assertions
# ---------------------------------------------------------------------------


def test_workbook_has_all_visible_sheets_in_canonical_order(
    sample_data: TaxOverviewData,
) -> None:
    wb = render_workbook(sample_data)
    assert wb.sheetnames == [
        "Übersicht",
        "Wertschriften",
        "Kauf_Verkauf",
        "Dividenden",
        "Zinsen",
        "Gebühren",
        "FX_Kurse",
    ]


def test_no_default_empty_sheet_remains(sample_data: TaxOverviewData) -> None:
    """openpyxl's 'Sheet' default must not leak into the final workbook."""
    wb = render_workbook(sample_data)
    assert "Sheet" not in wb.sheetnames


def test_no_ks36_sheet_in_non_preparer_mode(sample_data: TaxOverviewData) -> None:
    wb = render_workbook(sample_data)
    assert not any(name.startswith("_KS36") for name in wb.sheetnames)


# ---------------------------------------------------------------------------
# Übersicht
# ---------------------------------------------------------------------------


def test_uebersicht_title_mentions_year_and_broker(
    sample_data: TaxOverviewData,
) -> None:
    wb = render_workbook(sample_data)
    ws = wb["Übersicht"]
    title = ws.cell(row=1, column=1).value
    assert "2025" in title
    assert "IBKR" in title


def test_uebersicht_kpi_tiles_use_kpi_styles(
    sample_data: TaxOverviewData,
) -> None:
    wb = render_workbook(sample_data)
    ws = wb["Übersicht"]
    # First tile row is row 3 (after title + blank + tile grid).
    label_cell = ws.cell(row=3, column=1)
    value_cell = ws.cell(row=3, column=2)
    assert label_cell.style == StyleName.KPI_LABEL
    assert value_cell.style == StyleName.KPI_VALUE
    assert value_cell.value == D("100000")  # opening


def test_uebersicht_waterfall_rows_include_opening_and_closing(
    sample_data: TaxOverviewData,
) -> None:
    wb = render_workbook(sample_data)
    ws = wb["Übersicht"]
    cell_values = [
        ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)
    ]
    assert "Opening value" in cell_values
    assert "Closing value" in cell_values
    assert "Differenz (CHF)" in cell_values


# ---------------------------------------------------------------------------
# Wertschriften
# ---------------------------------------------------------------------------


def test_wertschriften_has_header_and_one_row_per_position(
    sample_data: TaxOverviewData,
) -> None:
    wb = render_workbook(sample_data)
    ws = wb["Wertschriften"]
    assert ws.cell(row=1, column=1).value == "ISIN"
    assert ws.cell(row=1, column=1).style == StyleName.HEADER
    # Header + 1 position
    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == "US0378331005"
    assert ws.cell(row=2, column=8).style == StyleName.BODY_CHF


# ---------------------------------------------------------------------------
# Kauf_Verkauf
# ---------------------------------------------------------------------------


def test_orders_sheet_lists_both_buy_and_sell(
    sample_data: TaxOverviewData,
) -> None:
    wb = render_workbook(sample_data)
    ws = wb["Kauf_Verkauf"]
    sides = [ws.cell(row=r, column=4).value for r in (2, 3)]
    assert set(sides) == {"BUY", "SELL"}


def test_orders_sheet_includes_realized_gain_table(
    sample_data: TaxOverviewData,
) -> None:
    wb = render_workbook(sample_data)
    ws = wb["Kauf_Verkauf"]
    # Scan for the realized-gain heading so the test doesn't depend on exact
    # row layout.
    headings = [
        ws.cell(row=r, column=1).value
        for r in range(1, ws.max_row + 1)
    ]
    assert "Realisierte Gewinne / Verluste" in headings


def test_orders_sheet_realized_pnl_value_matches_fixture(
    sample_data: TaxOverviewData,
) -> None:
    wb = render_workbook(sample_data)
    ws = wb["Kauf_Verkauf"]
    # Find the row with cost_per_share = 180.1 (the one lot close) and check
    # its P&L column (col 8).
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=6).value == D("180.1"):
            pnl = ws.cell(row=row, column=8).value
            # (209.9 - 180.1) * 10 = 298
            assert pnl == D("298.0")
            return
    pytest.fail("Did not find lot close row in Kauf_Verkauf sheet")


# ---------------------------------------------------------------------------
# Income (Dividenden / Zinsen)
# ---------------------------------------------------------------------------


def test_dividends_sheet_shows_gross_chf_with_chf_style(
    sample_data: TaxOverviewData,
) -> None:
    wb = render_workbook(sample_data)
    ws = wb["Dividenden"]
    cell = ws.cell(row=2, column=9)  # gross CHF column
    assert cell.value == D("22.50")
    assert cell.style == StyleName.BODY_CHF


def test_interest_sheet_separate_from_dividends(
    sample_data: TaxOverviewData,
) -> None:
    wb = render_workbook(sample_data)
    div_ws = wb["Dividenden"]
    int_ws = wb["Zinsen"]
    assert div_ws.cell(row=2, column=3).value == "AAPL"
    assert int_ws.cell(row=2, column=3).value == "CASH-USD"


# ---------------------------------------------------------------------------
# Gebühren / FX
# ---------------------------------------------------------------------------


def test_fees_sheet_has_fee_row(sample_data: TaxOverviewData) -> None:
    wb = render_workbook(sample_data)
    ws = wb["Gebühren"]
    assert ws.cell(row=2, column=2).value == "platform"
    assert ws.cell(row=2, column=6).value == D("9.00")


def test_fx_sheet_records_rate_and_source(sample_data: TaxOverviewData) -> None:
    wb = render_workbook(sample_data)
    ws = wb["FX_Kurse"]
    assert ws.cell(row=2, column=2).value == "USD"
    assert ws.cell(row=2, column=3).value == D("0.90")
    assert ws.cell(row=2, column=4).value == "kursliste"


# ---------------------------------------------------------------------------
# Totals helpers on TaxOverviewData
# ---------------------------------------------------------------------------


def test_data_aggregation_helpers(sample_data: TaxOverviewData) -> None:
    assert sample_data.total_dividends_chf() == D("22.50")
    assert sample_data.total_interest_chf() == D("10.80")
    assert sample_data.total_withholding_tax_chf() == D("3.38")
    # fees (9) + commissions (|-1| + |-1| = 2) = 11
    assert sample_data.total_fees_chf() == D("11.00")


def test_empty_sections_still_produce_sheet_with_header_only() -> None:
    minimal = TaxOverviewData(
        tax_year=2025,
        broker="ibkr",
        preparer_mode=False,
        opening_value_chf=D("0"),
        closing_value_chf=D("0"),
        waterfall=build_waterfall(opening=D("0"), closing=D("0")),
    )
    wb = render_workbook(minimal)
    assert "Wertschriften" in wb.sheetnames
    ws = wb["Wertschriften"]
    assert ws.max_row == 1  # header only
    assert ws.cell(row=1, column=1).value == "ISIN"
